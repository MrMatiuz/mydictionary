# import pandas as pd

# file = 'Words_copy.xlsx'

# df = pd.read_excel(file)

# new_id = df.iloc[-1].id
# new_word = 'qwerty'
# new_translation = 'кверти'
# df.append([new_id, new_word, new_translation])

# print(df)



from openpyxl import load_workbook

filename = 'Words_copy.xlsx'
wb = load_workbook(filename=filename)
sheet = wb['Sheet1']

last_row = sheet.max_row

new_id = last_row - 1
new_word = 'qwerty'
new_translation = 'кверти'

sheet.cell(row=last_row + 1, column=1, value=new_id)
sheet.cell(row=last_row + 1, column=2, value=new_word)
sheet.cell(row=last_row + 1, column=3, value=new_translation)

# sheet.append(new_row)
wb.save(filename)